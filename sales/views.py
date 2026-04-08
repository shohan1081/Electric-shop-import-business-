import openpyxl
from django.http import HttpResponse
from django.utils import timezone
from io import BytesIO
from django.shortcuts import get_object_or_404
from django.contrib.admin.views.decorators import staff_member_required
from .models import Customer, Sale, Payment
import decimal

@staff_member_required
def export_customer_history(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id)
    
    # 1. Create Workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'CustomerStatement'

    # Styles
    from openpyxl.styles import Font, Alignment, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    center_aligned = Alignment(horizontal="center")

    # 2. Customer Info Header
    worksheet.merge_cells('A1:E1')
    worksheet['A1'] = f"CUSTOMER STATEMENT: {customer.name.upper()}"
    worksheet['A1'].font = Font(size=14, bold=True)
    worksheet['A1'].alignment = center_aligned

    worksheet.append(["Phone:", customer.phone, "", "Current Due:", float(customer.total_due)])
    worksheet.append(["Address:", customer.address, "", "Date:", timezone.now().strftime('%Y-%m-%d')])
    worksheet.append([]) # Empty row

    # 3. Sales History Section
    worksheet.append(["SALES HISTORY"])
    worksheet['A5'].font = Font(bold=True)
    
    sales_headers = ["Date", "Product", "Quantity", "Unit Price", "Total Price"]
    worksheet.append(sales_headers)
    
    # Apply styling to headers
    for cell in worksheet[6]:
        cell.font = header_font
        cell.fill = header_fill

    sales = customer.sales.all().order_by('-sold_date')
    for sale in sales:
        worksheet.append([
            sale.sold_date.strftime('%Y-%m-%d %H:%M'),
            sale.product.name,
            sale.quantity_sold,
            float(sale.selling_price_at_that_time),
            float(sale.total_price)
        ])

    worksheet.append([]) # Empty row
    
    # 4. Payment History Section
    worksheet.append(["PAYMENT HISTORY"])
    # Dynamic positioning for Payments
    pay_row_start = worksheet.max_row
    worksheet.cell(row=pay_row_start, column=1).font = Font(bold=True)

    payment_headers = ["Date", "Amount Paid", "Note", "", ""]
    worksheet.append(payment_headers)
    for cell in worksheet[pay_row_start + 1]:
        cell.font = header_font
        cell.fill = header_fill

    payments = customer.payments.all().order_by('-payment_date')
    for pay in payments:
        worksheet.append([
            pay.payment_date.strftime('%Y-%m-%d %H:%M'),
            float(pay.amount_paid),
            pay.note or "",
            "", ""
        ])

    # Adjust Column Widths
    for col in ['A', 'B', 'C', 'D', 'E']:
        worksheet.column_dimensions[col].width = 20

    # 5. Save to BytesIO stream
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # 6. Build Response
    filename = f"statement_{customer.name}_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
